from __future__ import annotations

import csv
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(r"E:\sl_obsidian\converted_markdown\组织生活会")
OUT_MD = ROOT / "问题查摆与原因剖析专项审核报告.md"
OUT_CSV = ROOT / "problem-analysis-special-audit.csv"
OUT_XLSX = ROOT / "问题查摆与原因剖析专项审核.xlsx"


NON_PERSONAL = [
    "会议纪要", "整改清单", "问题清单", "查摆问题清单", "查摆问题及整改措施", "规范化要求", "工作汇报",
    "整改落实情况", "民主评议", "结果报告", "支部委员会", "支部对照", "支部班子",
    "班子发言", "支委班子", "党支部发言", "党支部材料", "支部材料", "参考模板", "模板",
]

SPECIFIC_MARKERS = [
    "例如", "比如", "具体", "实际", "项目", "客户", "群众", "网格", "支部",
    "岗位", "工作中", "学习", "走访", "调研", "投诉", "系统", "材料", "台账",
    "方案", "指标", "一线", "基层", "服务", "办理", "组织", "负责", "牵头",
]

PERSON_MARKERS = ["我", "本人", "个人", "自己", "牵头", "负责", "参与", "联系", "服务", "对接"]
EVENT_MARKERS = [
    "例如", "比如", "项目", "客户", "群众", "网格", "支部", "岗位", "工作中",
    "走访", "调研", "投诉", "会议", "材料", "台账", "方案", "系统", "学习",
    "办理", "推进", "落实", "服务", "支撑", "审核", "报送",
]

GENERIC_PHRASES = [
    "学习不够深入", "理论学习不够", "联系群众不够", "服务意识不强", "担当意识不强",
    "工作作风不够扎实", "创新意识不足", "先锋模范作用发挥不够", "廉洁自律意识不够",
    "宗旨意识有所淡化", "工作标准不高", "思想认识不到位",
]


def read_text(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="ignore")


def source_from_text(text: str) -> str:
    match = re.search(r"^Source: `(.+?)`", text, re.M)
    return match.group(1) if match else ""


def clean_text(text: str) -> str:
    lines = []
    for line in text.splitlines():
        if line.startswith("# "):
            continue
        if line.startswith("Source:"):
            continue
        if line.strip().startswith("|"):
            continue
        lines.append(line)
    return "\n".join(lines)


def compact_len(text: str) -> int:
    return len(re.sub(r"\s+", "", text))


def group_from_path(path: Path) -> str:
    rel = path.relative_to(ROOT)
    return rel.parts[0] if len(rel.parts) > 1 else "(root)"


def is_personal_material(path: Path, text: str) -> bool:
    name = path.name
    if any(k in name for k in NON_PERSONAL):
        return False
    if any(k in name for k in ["个人", "发言", "对照检查", "发言提纲"]):
        return True
    return "我" in text and "查摆" in text and "整改" in text


def infer_name(path: Path) -> str:
    stem = path.stem
    patterns = [
        r"个人发言材料[-_.：—]*([^（）()0-9vV]+)",
        r"个人对照检查材料[-_.：—]*([^（）()0-9vV]+)",
        r"发言提纲[-_.：—]*([^（）()0-9vV]+)",
        r"发言材料[-_.：—]*([^（）()0-9vV]+)",
        r"[-_—]([一-龥]{2,4})(?:\s|\(|（|$)",
        r"^([一-龥]{2,4})[+_.-]",
    ]
    for pattern in patterns:
        match = re.search(pattern, stem)
        if match:
            value = re.sub(
                r"(办公室|客服部|支部书记|宣传委员|组织委员|纪检委员|青年委员|党支部|人力资源部党支部|市场部|综服中心|数字化与稽核中心)",
                "",
                match.group(1),
            )
            value = value.strip(" -_—.（）()")
            if 2 <= len(value) <= 4:
                return value
    return stem


def find_sections(text: str) -> tuple[str, str, str]:
    body = clean_text(text)
    problem_match = re.search(r"(查摆问题|问题查摆|存在的?问题|突出问题|对照检查)", body)
    cause_match = re.search(r"(原因剖析|根源剖析|产生问题的原因|问题根源|原因分析)", body)
    rect_match = re.search(r"(整改措施|努力方向|下一步|整改方向|改进措施)", body)

    problem_start = problem_match.start() if problem_match else -1
    cause_start = cause_match.start() if cause_match else -1
    rect_start = rect_match.start() if rect_match else -1

    if problem_start >= 0 and cause_start >= 0 and cause_start > problem_start:
        problem_text = body[problem_start:cause_start]
    elif problem_start >= 0 and rect_start > problem_start:
        problem_text = body[problem_start:rect_start]
    elif problem_start >= 0:
        problem_text = body[problem_start:]
    else:
        problem_text = ""

    if cause_start >= 0 and rect_start > cause_start:
        cause_text = body[cause_start:rect_start]
    elif cause_start >= 0:
        cause_text = body[cause_start:]
    else:
        cause_text = ""

    combined = (problem_text + "\n" + cause_text).strip()
    return problem_text, cause_text, combined


def specific_score(problem_text: str) -> tuple[int, list[str]]:
    hits = []
    for marker in SPECIFIC_MARKERS:
        if marker in problem_text:
            hits.append(marker)
    person_hit = any(marker in problem_text for marker in PERSON_MARKERS)
    event_hit = any(marker in problem_text for marker in EVENT_MARKERS)
    number_hit = bool(re.search(r"\d+|[一二三四五六七八九十]+(次|项|个|篇|月|年|小时|天)", problem_text))
    example_hit = "例如" in problem_text or "比如" in problem_text
    generic_count = sum(problem_text.count(p) for p in GENERIC_PHRASES)

    score = 0
    score += min(len(set(hits)), 8)
    score += 3 if person_hit else 0
    score += 3 if event_hit else 0
    score += 2 if number_hit else 0
    score += 2 if example_hit else 0
    score -= min(generic_count, 5)
    return score, sorted(set(hits))


def evaluate(path: Path) -> dict[str, object]:
    text = read_text(path)
    body = clean_text(text)
    problem_text, cause_text, combined = find_sections(text)
    total_chars = compact_len(body)
    problem_chars = compact_len(problem_text)
    cause_chars = compact_len(cause_text)
    combined_chars = compact_len(combined)
    ratio = combined_chars / total_chars if total_chars else 0
    score, evidence = specific_score(problem_text)

    ratio_ok = ratio >= 0.5
    has_problem = problem_chars >= 500
    has_cause = cause_chars >= 300
    concrete_ok = score >= 9 and has_problem

    issues = []
    if not ratio_ok:
        issues.append(f"问题查摆+原因剖析占比不足1/2（当前约{ratio:.1%}）")
    if not has_problem:
        issues.append("问题查摆篇幅偏少或未识别到完整查摆部分")
    if not has_cause:
        issues.append("原因剖析篇幅偏少或未识别到完整剖析部分")
    if not concrete_ok:
        issues.append("问题查摆偏简单，见人见事不足")

    if not ratio_ok and not concrete_ok:
        level = "不达标"
    elif issues:
        level = "需修改"
    else:
        level = "基本达标"

    return {
        "level": level,
        "group": group_from_path(path),
        "name": infer_name(path),
        "total_chars": total_chars,
        "problem_chars": problem_chars,
        "cause_chars": cause_chars,
        "problem_cause_chars": combined_chars,
        "problem_cause_ratio": ratio,
        "specific_score": score,
        "specific_evidence": "、".join(evidence[:12]),
        "issues": "；".join(issues),
        "file": str(path),
        "source": source_from_text(text),
    }


def candidates() -> list[Path]:
    result = []
    for path in ROOT.rglob("*.md"):
        if path.name in {"index.md", "conversion-summary.md", "个人对照检查材料审核报告.md", OUT_MD.name}:
            continue
        text = read_text(path)
        if is_personal_material(path, text):
            result.append(path)
    return sorted(result)


def write_csv(rows: list[dict[str, object]]) -> None:
    fields = [
        "level", "group", "name", "total_chars", "problem_chars", "cause_chars",
        "problem_cause_chars", "problem_cause_ratio", "specific_score",
        "specific_evidence", "issues", "file", "source",
    ]
    with OUT_CSV.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(rows: list[dict[str, object]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "专项审核"
    headers = [
        "结论", "支部/单位", "人员/材料", "全文字数", "问题查摆字数", "原因剖析字数",
        "查摆+剖析字数", "占比", "具体性评分", "具体性依据", "主要问题", "Markdown文件", "源文件",
    ]
    ws.append(headers)
    for row in rows:
        ws.append([
            row["level"], row["group"], row["name"], row["total_chars"], row["problem_chars"],
            row["cause_chars"], row["problem_cause_chars"], row["problem_cause_ratio"],
            row["specific_score"], row["specific_evidence"], row["issues"], row["file"], row["source"],
        ])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        if row[0].value == "不达标":
            row[0].fill = PatternFill("solid", fgColor="F4CCCC")
        elif row[0].value == "需修改":
            row[0].fill = PatternFill("solid", fgColor="FFF2CC")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = [10, 14, 28, 10, 12, 12, 14, 10, 12, 32, 48, 56, 56]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row_idx, 8).number_format = "0.0%"
    wb.save(OUT_XLSX)


def write_report(rows: list[dict[str, object]]) -> None:
    bad = [r for r in rows if r["level"] == "不达标"]
    needs = [r for r in rows if r["level"] == "需修改"]
    ok = [r for r in rows if r["level"] == "基本达标"]
    lines = [
        "# 问题查摆与原因剖析专项审核报告",
        "",
        "专项检查口径：",
        "",
        "- 问题查摆和原因剖析合计不少于全文 1/2。",
        "- 问题查摆不能过于简单，要能体现具体岗位、具体事项、具体表现，做到见人见事。",
        "",
        f"审核范围：`{ROOT}` 下已转换的个人发言/个人对照检查材料。",
        "",
        "## 统计",
        "",
        f"- 纳入审核：{len(rows)} 份",
        f"- 不达标：{len(bad)} 份",
        f"- 需修改：{len(needs)} 份",
        f"- 基本达标：{len(ok)} 份",
        "",
        "## 不达标名单",
        "",
        "| 支部/单位 | 人员/材料 | 查摆+剖析占比 | 具体性评分 | 主要问题 | 文件 |",
        "| --- | --- | ---: | ---: | --- | --- |",
    ]
    for r in bad:
        lines.append(
            f"| {r['group']} | {r['name']} | {r['problem_cause_ratio']:.1%} | {r['specific_score']} | {r['issues']} | `{r['file']}` |"
        )
    lines.extend([
        "",
        "## 需修改名单",
        "",
        "| 支部/单位 | 人员/材料 | 查摆+剖析占比 | 具体性评分 | 主要问题 | 文件 |",
        "| --- | --- | ---: | ---: | --- | --- |",
    ])
    for r in needs:
        lines.append(
            f"| {r['group']} | {r['name']} | {r['problem_cause_ratio']:.1%} | {r['specific_score']} | {r['issues']} | `{r['file']}` |"
        )
    lines.extend(["", "## 基本达标名单", ""])
    for r in ok:
        lines.append(f"- {r['group']}：{r['name']}（占比 {r['problem_cause_ratio']:.1%}，具体性评分 {r['specific_score']}）")
    OUT_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    rows = [evaluate(path) for path in candidates()]
    rows.sort(key=lambda r: ({"不达标": 0, "需修改": 1, "基本达标": 2}[str(r["level"])], str(r["group"]), str(r["name"])))
    write_csv(rows)
    write_xlsx(rows)
    write_report(rows)
    print(f"audited={len(rows)} report={OUT_MD} excel={OUT_XLSX}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
