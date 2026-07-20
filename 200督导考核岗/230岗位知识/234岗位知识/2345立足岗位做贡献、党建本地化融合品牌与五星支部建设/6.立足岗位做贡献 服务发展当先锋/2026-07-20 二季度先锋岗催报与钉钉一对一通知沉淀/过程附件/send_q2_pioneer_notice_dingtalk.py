from __future__ import annotations

import argparse
import json
import subprocess
import sys
import uuid
from pathlib import Path
from typing import Any

import openpyxl


SOURCE = r"D:\work\2026\2026年专职党务人员信息统计表.xlsx"
RESULT = "二季度党员先锋岗申报评选材料提交通知_钉钉发送结果.xlsx"
TITLE = "关于二季度党员先锋岗申报及评选材料提交通知"
NOTICE = """关于二季度“立足岗位做贡献、服务发展当先锋”党员先锋岗申报及评选材料提交通知

各支部组织委员：
      现请准备2026年二季度党员先锋岗申报及评选材料提交。
      请各支部根据考核细则，对二季度党员先锋岗申报人员进行打分，申报人数不超过本支部党员数的20%（向上取整）。
      申报表须经支部书记审核签字，同时报送签字盖章版和可编辑版。
      各单位打分结果须加盖支部章，并将盖章结果图片及Excel表于7月22日18:00前通过钉钉发党建工作部。"""


def run_dws(args: list[str]) -> tuple[bool, Any]:
    cmd = ["dws.cmd", *args, "--format", "json"]
    proc = subprocess.run(cmd, capture_output=True, text=True, encoding="utf-8", errors="replace")
    output = (proc.stdout or proc.stderr or "").strip()
    try:
        payload = json.loads(output)
    except Exception:
        payload = output
    return proc.returncode == 0, payload


def walk(obj: Any):
    if isinstance(obj, dict):
        for key, value in obj.items():
            yield key, value
            yield from walk(value)
    elif isinstance(obj, list):
        for item in obj:
            yield from walk(item)


def first_field(payload: Any, names: set[str]) -> str:
    for key, value in walk(payload):
        if key in names and value:
            return str(value)
    return ""


def read_people(path: str) -> list[dict[str, str]]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    people: list[dict[str, str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        dept, name, mobile = [(str(v).strip() if v is not None else "") for v in row[:3]]
        if dept and name and mobile:
            people.append({"dept": dept, "name": name, "mobile": mobile})
    return people


def resolve_user(mobile: str) -> tuple[str, str]:
    ok, payload = run_dws(["contact", "user", "search-mobile", "--mobile", mobile])
    if not ok:
        return "", json.dumps(payload, ensure_ascii=False)[:800]
    user_id = first_field(payload, {"userId", "userid", "orgUserId", "staffId"})
    return user_id, json.dumps(payload, ensure_ascii=False)[:800]


def send_to_user(user_id: str, mobile: str) -> tuple[str, str]:
    msg_uuid = str(uuid.uuid5(uuid.NAMESPACE_DNS, f"q2-party-pioneer-notice-20260720-{mobile}"))
    ok, payload = run_dws([
        "chat",
        "message",
        "send",
        "--user",
        user_id,
        "--title",
        TITLE,
        "--text",
        NOTICE,
        "--uuid",
        msg_uuid,
    ])
    status = "sent" if ok else "failed"
    detail = json.dumps(payload, ensure_ascii=False)[:800] if not isinstance(payload, str) else payload[:800]
    return status, detail


def write_result(rows: list[dict[str, str]], path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "发送结果"
    headers = ["部门", "姓名", "联系电话", "userId", "匹配状态", "发送状态", "匹配详情", "发送详情"]
    ws.append(headers)
    for row in rows:
        ws.append([
            row["dept"],
            row["name"],
            row["mobile"],
            row.get("user_id", ""),
            row.get("resolve_status", ""),
            row.get("send_status", ""),
            row.get("resolve_detail", ""),
            row.get("send_detail", ""),
        ])
    widths = [16, 12, 16, 28, 12, 12, 70, 70]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor="D9EAF7")
    wb.save(path)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", default=SOURCE)
    parser.add_argument("--out", default=RESULT)
    parser.add_argument("--send", action="store_true")
    parser.add_argument("--yes", action="store_true")
    args = parser.parse_args()

    if args.send and not args.yes:
        print("Refusing to send without --yes.", file=sys.stderr)
        return 2

    rows: list[dict[str, str]] = []
    for person in read_people(args.source):
        user_id, resolve_detail = resolve_user(person["mobile"])
        person["user_id"] = user_id
        person["resolve_status"] = "ok" if user_id else "failed"
        person["resolve_detail"] = resolve_detail
        if args.send and user_id:
            send_status, send_detail = send_to_user(user_id, person["mobile"])
        else:
            send_status, send_detail = ("not_sent", "")
        person["send_status"] = send_status
        person["send_detail"] = send_detail
        rows.append(person)

    out = str(Path(args.out).resolve())
    write_result(rows, out)
    print(out)
    print(json.dumps({
        "total": len(rows),
        "resolved": sum(1 for r in rows if r.get("resolve_status") == "ok"),
        "sent": sum(1 for r in rows if r.get("send_status") == "sent"),
        "failed_resolve": [f"{r['dept']}-{r['name']}" for r in rows if r.get("resolve_status") != "ok"],
        "failed_send": [f"{r['dept']}-{r['name']}" for r in rows if r.get("send_status") == "failed"],
    }, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
