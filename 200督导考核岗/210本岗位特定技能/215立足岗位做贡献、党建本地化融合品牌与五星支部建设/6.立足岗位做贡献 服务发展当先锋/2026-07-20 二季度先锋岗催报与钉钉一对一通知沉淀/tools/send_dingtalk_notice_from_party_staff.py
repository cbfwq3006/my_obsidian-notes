from __future__ import annotations

import argparse
import json
import subprocess
import sys
import uuid
from pathlib import Path
from typing import Any

import openpyxl


def run_dws(args: list[str]) -> tuple[bool, Any]:
    cmd = ["dws.cmd", *args, "--format", "json"]
    proc = subprocess.run(cmd, capture_output=True, text=True, encoding="utf-8", errors="replace")
    output = (proc.stdout or proc.stderr or "").strip()
    try:
        payload = json.loads(output)
    except Exception:
        payload = output
    return proc.returncode == 0, payload


def walk(obj: Any):
    if isinstance(obj, dict):
        for key, value in obj.items():
            yield key, value
            yield from walk(value)
    elif isinstance(obj, list):
        for item in obj:
            yield from walk(item)


def first_field(payload: Any, names: set[str]) -> str:
    for key, value in walk(payload):
        if key in names and value:
            return str(value)
    return ""


def read_people(path: Path) -> list[dict[str, str]]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    header = [str(v).strip() if v is not None else "" for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    index = {name: i for i, name in enumerate(header)}
    people: list[dict[str, str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        values = [str(v).strip() if v is not None else "" for v in row]
        dept = values[index.get("部门", 0)] if len(values) > index.get("部门", 0) else ""
        name = values[index.get("姓名", 1)] if len(values) > index.get("姓名", 1) else ""
        mobile = values[index.get("联系电话", 2)] if len(values) > index.get("联系电话", 2) else ""
        user_id = values[index["userId"]] if "userId" in index and len(values) > index["userId"] else ""
        open_dingtalk_id = values[index["openDingTalkId"]] if "openDingTalkId" in index and len(values) > index["openDingTalkId"] else ""
        if dept and name and mobile:
            people.append({
                "dept": dept,
                "name": name,
                "mobile": mobile,
                "user_id": user_id,
                "open_dingtalk_id": open_dingtalk_id,
            })
    return people


def resolve_by_mobile(mobile: str) -> tuple[str, str]:
    ok, payload = run_dws(["contact", "user", "search-mobile", "--mobile", mobile])
    detail = json.dumps(payload, ensure_ascii=False)[:800] if not isinstance(payload, str) else payload[:800]
    if not ok:
        return "", detail
    return first_field(payload, {"userId", "userid", "orgUserId", "staffId"}), detail


def send_notice(target_flag: str, target_value: str, title: str, text: str, stable_key: str, send_title: bool) -> tuple[str, str]:
    message_uuid = str(uuid.uuid5(uuid.NAMESPACE_DNS, stable_key))
    args = [
        "chat",
        "message",
        "send",
        target_flag,
        target_value,
        "--text",
        text,
        "--uuid",
        message_uuid,
    ]
    if send_title:
        args.extend(["--title", title])
    ok, payload = run_dws(args)
    status = "sent" if ok else "failed"
    detail = json.dumps(payload, ensure_ascii=False)[:800] if not isinstance(payload, str) else payload[:800]
    return status, detail


def write_result(rows: list[dict[str, str]], out: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "发送结果"
    ws.append(["部门", "姓名", "联系电话", "userId", "openDingTalkId", "匹配状态", "发送状态", "匹配详情", "发送详情"])
    for row in rows:
        ws.append([
            row["dept"],
            row["name"],
            row["mobile"],
            row.get("user_id", ""),
            row.get("open_dingtalk_id", ""),
            row.get("resolve_status", ""),
            row.get("send_status", ""),
            row.get("resolve_detail", ""),
            row.get("send_detail", ""),
        ])
    for idx, width in enumerate([16, 12, 16, 28, 32, 16, 12, 70, 70], start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor="D9EAF7")
    wb.save(out)


def main() -> int:
    parser = argparse.ArgumentParser(description="Send DingTalk one-to-one notices from party staff xlsx.")
    parser.add_argument("--source", required=True, help="Excel with columns: 部门, 姓名, 联系电话.")
    parser.add_argument("--title", required=True, help="DingTalk message title.")
    parser.add_argument("--text-file", required=True, help="UTF-8 text file containing notice body.")
    parser.add_argument("--out", default="钉钉发送结果.xlsx", help="Output xlsx result file.")
    parser.add_argument("--send", action="store_true", help="Actually send messages. Without this flag, only resolves users.")
    parser.add_argument("--send-title", action="store_true", help="Also pass --title to DWS. Default is false to avoid title-only display issues.")
    parser.add_argument("--yes", action="store_true", help="Required with --send.")
    args = parser.parse_args()

    if args.send and not args.yes:
        print("Refusing to send without --yes.", file=sys.stderr)
        return 2

    source = Path(args.source)
    text = Path(args.text_file).read_text(encoding="utf-8")
    rows: list[dict[str, str]] = []
    for person in read_people(source):
        resolve_detail = ""
        if person.get("user_id"):
            person["resolve_status"] = "ok_manual_userId"
            resolve_detail = "使用人员表中预置 userId，未再按手机号查询。"
        elif person.get("open_dingtalk_id"):
            person["resolve_status"] = "ok_manual_openDingTalkId"
            resolve_detail = "使用人员表中预置 openDingTalkId，未再按手机号查询。"
        else:
            user_id, resolve_detail = resolve_by_mobile(person["mobile"])
            person["user_id"] = user_id
            person["resolve_status"] = "ok" if user_id else "failed"
        person["resolve_detail"] = resolve_detail
        if args.send and (person.get("user_id") or person.get("open_dingtalk_id")):
            key = f"{args.title}-{person['mobile']}-{text[:30]}"
            if person.get("user_id"):
                person["send_status"], person["send_detail"] = send_notice("--user", person["user_id"], args.title, text, key, args.send_title)
            else:
                person["send_status"], person["send_detail"] = send_notice("--open-dingtalk-id", person["open_dingtalk_id"], args.title, text, key, args.send_title)
        else:
            person["send_status"] = "not_sent"
            person["send_detail"] = ""
        rows.append(person)

    out = Path(args.out).resolve()
    write_result(rows, out)
    print(out)
    print(json.dumps({
        "total": len(rows),
        "resolved": sum(1 for r in rows if r.get("resolve_status") == "ok"),
        "sent": sum(1 for r in rows if r.get("send_status") == "sent"),
        "failed_resolve": [f"{r['dept']}-{r['name']}" for r in rows if r.get("resolve_status") != "ok"],
        "failed_send": [f"{r['dept']}-{r['name']}" for r in rows if r.get("send_status") == "failed"],
    }, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
