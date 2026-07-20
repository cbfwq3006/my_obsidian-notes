from __future__ import annotations

import argparse
import json
import subprocess
import sys
import uuid
from pathlib import Path
from typing import Any

import openpyxl


DEFAULT_SOURCE = r"D:\work\2026\2026年专职党务人员信息统计表.xlsx"
DEFAULT_PREVIEW = "党员先锋岗二季度催报及三季度指引_钉钉待发送预览.xlsx"
TITLE = "二季度先锋岗催报及三季度活动指引"


def build_message(name: str, dept: str) -> str:
    return (
        f"{name}同志好，请协助落实{dept}党支部“立足岗位做贡献、服务发展当先锋”主题实践活动两项工作：\n\n"
        "1. 报送二季度结果。请对照《党员先锋岗考核细则》和负面清单，报送2026年二季度（4-6月）党员先锋岗评选推荐结果；本季度无推荐对象的，也请反馈“二季度无推荐”。\n"
        "2. 启动三季度工作。请支部围绕三季度（7-9月）中心任务和短板指标，确定党员先锋岗创建对象、党员先锋队或党员领题攻坚事项，做到季初定题、季中推题、季末评题。\n"
        "3. 同步建立过程台账。三季度不等到季末集中补材料，每月至少记录一次进展、一次数据变化、一次典型做法。\n\n"
        "请于7月23日（周四）18:00前反馈：二季度党员先锋岗推荐情况或无推荐说明、三季度攻坚事项和创建对象、需要公司层面协调的问题。"
        "相关表格按原方案附件执行，操作方法可参考《支部一眼看懂操作指引》。收到请回复。"
    )


def read_people(path: Path) -> list[dict[str, str]]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    people: list[dict[str, str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        dept, name, mobile = [(str(v).strip() if v is not None else "") for v in row[:3]]
        if dept and name and mobile:
            people.append({"dept": dept, "name": name, "mobile": mobile})
    return people


def run_dws(args: list[str]) -> tuple[bool, dict[str, Any] | str]:
    cmd = ["dws.cmd", *args, "--format", "json"]
    proc = subprocess.run(cmd, capture_output=True, text=True, encoding="utf-8", errors="replace")
    text = (proc.stdout or proc.stderr or "").strip()
    try:
        payload: dict[str, Any] | str = json.loads(text)
    except Exception:
        payload = text
    return proc.returncode == 0, payload


def walk_values(obj: Any):
    if isinstance(obj, dict):
        for key, value in obj.items():
            yield key, value
            yield from walk_values(value)
    elif isinstance(obj, list):
        for item in obj:
            yield from walk_values(item)


def first_field(payload: Any, names: set[str]) -> str:
    for key, value in walk_values(payload):
        if key in names and value:
            return str(value)
    return ""


def resolve_user(mobile: str) -> dict[str, str]:
    ok, payload = run_dws(["contact", "user", "search-mobile", "--mobile", mobile])
    if not ok:
        return {
            "user_id": "",
            "open_dingtalk_id": "",
            "resolve_status": "failed",
            "resolve_detail": json.dumps(payload, ensure_ascii=False)[:500],
        }
    return {
        "user_id": first_field(payload, {"userId", "userid", "orgUserId", "staffId"}),
        "open_dingtalk_id": first_field(payload, {"openDingTalkId", "openDingtalkId"}),
        "resolve_status": "ok",
        "resolve_detail": json.dumps(payload, ensure_ascii=False)[:500],
    }


def write_preview(path: Path, rows: list[dict[str, str]]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "待发送预览"
    headers = ["部门", "姓名", "联系电话", "userId", "openDingTalkId", "匹配状态", "发送状态", "消息标题", "消息正文", "匹配详情"]
    ws.append(headers)
    for row in rows:
        ws.append([
            row["dept"],
            row["name"],
            row["mobile"],
            row.get("user_id", ""),
            row.get("open_dingtalk_id", ""),
            row.get("resolve_status", "not_resolved"),
            row.get("send_status", "not_sent"),
            TITLE,
            row["message"],
            row.get("resolve_detail", ""),
        ])
    widths = [14, 12, 16, 28, 28, 14, 14, 28, 90, 60]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor="D9EAF7")
    wb.save(path)


def send_message(row: dict[str, str]) -> dict[str, str]:
    user_id = row.get("user_id", "")
    if not user_id:
        return {"send_status": "skipped_no_user_id", "send_detail": ""}
    msg_uuid = str(uuid.uuid5(uuid.NAMESPACE_DNS, f"party-pioneer-2026q2-q3-{row['mobile']}"))
    ok, payload = run_dws([
        "chat",
        "message",
        "send",
        "--user",
        user_id,
        "--title",
        TITLE,
        "--text",
        row["message"],
        "--uuid",
        msg_uuid,
    ])
    return {
        "send_status": "sent" if ok else "failed",
        "send_detail": json.dumps(payload, ensure_ascii=False)[:500] if not isinstance(payload, str) else payload[:500],
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Prepare or send one-to-one DingTalk notices for party pioneer activity.")
    parser.add_argument("--source", default=DEFAULT_SOURCE, help="Input party staff xlsx.")
    parser.add_argument("--out", default=DEFAULT_PREVIEW, help="Output preview xlsx.")
    parser.add_argument("--resolve", action="store_true", help="Resolve DingTalk userId by mobile through dws.")
    parser.add_argument("--send", action="store_true", help="Actually send messages through dws chat message send.")
    parser.add_argument("--yes", action="store_true", help="Required together with --send.")
    args = parser.parse_args()

    if args.send and not args.yes:
        print("Refusing to send: add --yes after checking the preview.", file=sys.stderr)
        return 2

    rows: list[dict[str, str]] = []
    for person in read_people(Path(args.source)):
        person["message"] = build_message(person["name"], person["dept"])
        if args.resolve or args.send:
            person.update(resolve_user(person["mobile"]))
        if args.send:
            person.update(send_message(person))
        rows.append(person)

    out_path = Path(args.out).resolve()
    write_preview(out_path, rows)
    print(out_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
